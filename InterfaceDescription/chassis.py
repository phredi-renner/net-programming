############################################################################
# Uses input from spreadsheet to create a textfile to add port
# description and basic port configuration.
# This can then be pasted into the appropriate configlet in Cloud Vision
# Or directly into a switch.  This version works on a Chassis based switch.
#
# Created by: Fred Renner
# Date: 3/10/2023
# V1.1 - added comments and info to reference proper CV configlet
# V1.2 - 4/25/23 - adding phone trunk untagged
# V1.3 - 5/9/23 - modificaitons to input spreadsheet formatting
############################################################################

import openpyxl

# Gather some info
work = input('Enter spreadsheet filename: ')
wkbk = 'input/'+work+'.xlsx'

# Open the spreadsheet where we have the data
wb = openpyxl.load_workbook(wkbk)

# Here we will look at each worksheet and gather the information
for sheet in wb.sheetnames:
    # define stuff we will work with
    page = wb[sheet]
    header = page['A1'].value
    sw_hostname = page['B3'].value
   #outputfile = 'output/'+header +'.ios'
    idf = page['E2'].value
    bldg = page['F2'].value
    sw_ip = page['G2'].value
    d_vlan = page['H2'].value
    v_vlan = page['I2'].value
    s_vlan = page['J2'].value
    sp_vlan = page['K2'].value
    outputfile = 'output/cus_interface_'+ str(sw_hostname) +'_' + str(sw_ip) + '.ios'
    with open(outputfile, 'w') as output:
    #    print(f"Page equals: {page}")
        # Create a header so we know what we are getting
        output.write('!============================\n')
        output.write('!CV filename:cus_interface_' + str(sw_hostname) + '_' + str(sw_ip) + '\n')
        output.write('!\n')
        output.write('! Paste everything below in the configlet\n')
        output.write('!-------\n')
        output.write('!CusConfig:interface_' + str(sw_hostname) + '_' + str(sw_ip) + '\n')
        output.write('!============================\n')
        output.write('!' + str(sw_hostname) + '\n')
        output.write('!' + str(header) + ' - ' + str(sheet) + '\n')
        output.write('!============================\n')
        output.write('\n')
        # Look at each row to create the switchport interface config
        # This would be cool if it was based on a jinja template :-)
        for i in range(3, page.max_row+1, 1):
            # If it's a wireless AP we use a different port configuration
            if 'AP' in str(page.cell(row=i, column=1).value):
                output.write(f'interface Ethernet' + str(page.cell(row=i, column=3).value) + '/' + str(page.cell(row=i, column=4).value) + '\n')
                output.write(f' description '+ str(bldg) + '-' + str(idf) + '-' + str(page.cell(row=i, column=1).value) + '\n')
                output.write(' switchport access vlan ' + str(d_vlan) + '\n')
                output.write(' no poe disable\n')
                output.write(' no shutdown\n')
                output.write('!\n')
            # if it's a camera then we have a different vlan
            elif 'CAM' in str(page.cell(row=i, column=1).value):
                output.write(f'interface Ethernet' + str(page.cell(row=i, column=3).value) + '/' + str(page.cell(row=i, column=4).value) + '\n')
                output.write(f' description '+ str(bldg) + '-' + str(idf) + '-' + str(page.cell(row=i, column=1).value) + '\n')
                output.write(' switchport access vlan ' + str(s_vlan) + '\n')
                output.write(' no poe disable\n')
                output.write(' no shutdown\n')
                output.write('!\n')
            elif 'SP' in str(page.cell(row=i, column=1).value):
                output.write(f'interface Ethernet' + str(page.cell(row=i, column=3).value) + '/' + str(page.cell(row=i, column=4).value) + '\n')
                output.write(f' description '+ str(bldg) + '-' + str(idf) + '-' + str(page.cell(row=i, column=1).value) + '\n')
                output.write(' switchport access vlan ' + str(sp_vlan) + '\n')
                output.write(' no poe disable\n')
                output.write(' no shutdown\n')
                output.write('!\n')
            else:
                # Default port configuration
                output.write(f'interface Ethernet' + str(page.cell(row=i, column=3).value) + '/' + str(page.cell(row=i, column=4).value) + '\n')
                output.write(f' description '+ str(bldg) + '-' + str(idf) + '-' + str(page.cell(row=i, column=1).value) + '\n')
            #    output.write(' switchport access vlan ' + str(d_vlan) + '\n')
                output.write(' switchport trunk native vlan ' + str(d_vlan) + '\n')
                output.write(' switchport phone vlan ' + str(v_vlan) + '\n')
                output.write(' switchport phone trunk untagged\n')
                output.write(' switchport mode trunk phone\n')
                output.write(' no poe disable\n')
                output.write(' no shutdown\n')
                output.write('!\n')
        print(f"... wrote {outputfile}")